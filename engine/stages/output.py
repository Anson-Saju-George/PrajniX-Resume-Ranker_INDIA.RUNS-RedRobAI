"""Submission rows, strict format guards, and deterministic CSV writing."""

from __future__ import annotations

import csv
import math
import re
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


HEADER = ("candidate_id", "rank", "score", "reasoning")
CANDIDATE_ID_PATTERN = re.compile(r"^CAND_[0-9]{7}$")


@dataclass(frozen=True, slots=True)
class SubmissionRow:
    candidate_id: str
    rank: int
    score: float
    reasoning: str


def validate_submission_rows(
    rows: Iterable[SubmissionRow], valid_candidate_ids: set[str] | frozenset[str]
) -> list[SubmissionRow]:
    """Enforce every repository rule before a byte is written."""

    materialized = list(rows)
    if len(materialized) != 100:
        raise ValueError(f"Submission must contain exactly 100 rows; got {len(materialized)}")
    if [row.rank for row in materialized] != list(range(1, 101)):
        raise ValueError("Ranks must appear in ascending order from 1 through 100")

    ids = [row.candidate_id for row in materialized]
    if len(ids) != len(set(ids)):
        raise ValueError("Submission candidate IDs must be unique")
    if any(not CANDIDATE_ID_PATTERN.fullmatch(candidate_id) for candidate_id in ids):
        raise ValueError("Submission contains a malformed candidate ID")
    if any(candidate_id not in valid_candidate_ids for candidate_id in ids):
        raise ValueError("Submission contains an ID absent from the source dataset")

    for index, row in enumerate(materialized):
        if not math.isfinite(row.score):
            raise ValueError(f"Rank {row.rank} has a non-finite score")
        if not row.reasoning.strip():
            raise ValueError(f"Rank {row.rank} has empty reasoning")
        if index == 0:
            continue
        previous = materialized[index - 1]
        if previous.score < row.score:
            raise ValueError("Scores must be non-increasing by rank")
        if previous.score == row.score and previous.candidate_id > row.candidate_id:
            raise ValueError("Equal scores must tie-break by candidate_id ascending")
    return materialized


def write_submission(path: str | Path, rows: Iterable[SubmissionRow]) -> None:
    """Write a guarded UTF-8 CSV with the exact required header order."""

    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.writer(output_file)
        writer.writerow(HEADER)
        for row in rows:
            writer.writerow((row.candidate_id, row.rank, f"{row.score:.8f}", row.reasoning))


def _read_submission_csv(path: Path) -> list[list[str]]:
    """Read the finished CSV without coercing identifiers or numeric text."""

    with path.open("r", encoding="utf-8", newline="") as source:
        rows = list(csv.reader(source))
    if not rows or tuple(rows[0]) != HEADER:
        raise ValueError(f"CSV header must be exactly {HEADER}")
    if len(rows) != 101:
        raise ValueError(f"XLSX conversion requires 100 data rows; found {len(rows) - 1}")
    return rows


def write_xlsx_from_csv(csv_path: str | Path) -> Path:
    """Create a typed, single-sheet XLSX from the just-written submission CSV."""

    source = Path(csv_path)
    rows = _read_submission_csv(source)
    destination = source.with_suffix(".xlsx")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Submission"
    worksheet.freeze_panes = "A2"
    worksheet.append(list(HEADER))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for candidate_id, rank, score, reasoning in rows[1:]:
        worksheet.append(
            (candidate_id, int(rank), Decimal(score), reasoning)
        )
        row_number = worksheet.max_row
        worksheet.cell(row_number, 1).number_format = "@"
        worksheet.cell(row_number, 3).number_format = "0.00000000"
        worksheet.cell(row_number, 4).alignment = Alignment(wrap_text=True)

    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 100
    worksheet.auto_filter.ref = "A1:D101"
    workbook.save(destination)
    return destination


def verify_xlsx_against_csv(
    csv_path: str | Path, xlsx_path: str | Path
) -> dict[str, Any]:
    """Reload the XLSX and prove value/order/type fidelity against its source CSV."""

    csv_rows = _read_submission_csv(Path(csv_path))
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    if workbook.sheetnames != ["Submission"]:
        raise ValueError("XLSX must contain one sheet named Submission")
    worksheet = workbook["Submission"]
    xlsx_rows = list(worksheet.iter_rows(min_row=1, max_col=4))

    header = [cell.value for cell in xlsx_rows[0]]
    mismatches: list[int] = []
    type_failures: list[str] = []
    preview: list[dict[str, Any]] = []
    loaded_scores: list[Decimal] = []
    loaded_ranks: list[int] = []

    for index, (csv_row, cells) in enumerate(zip(csv_rows[1:], xlsx_rows[1:]), 1):
        candidate_id, rank, score, reasoning = csv_row
        loaded = [cell.value for cell in cells]
        score_matches = Decimal(str(loaded[2])) == Decimal(score)
        values_match = (
            loaded[0] == candidate_id
            and loaded[1] == int(rank)
            and score_matches
            and loaded[3] == reasoning
        )
        if not values_match:
            mismatches.append(index)
        if cells[0].data_type != "s":
            type_failures.append(f"A{index + 1}")
        if cells[1].data_type != "n" or not isinstance(loaded[1], int):
            type_failures.append(f"B{index + 1}")
        if cells[2].data_type != "n":
            type_failures.append(f"C{index + 1}")
        if cells[3].data_type != "s":
            type_failures.append(f"D{index + 1}")
        loaded_ranks.append(int(loaded[1]))
        loaded_scores.append(Decimal(str(loaded[2])))
        preview.append(
            {
                "candidate_id": loaded[0],
                "rank": loaded[1],
                "score": score,
                "reasoning": loaded[3],
                "types": ("TEXT", "INTEGER", "NUMERIC", "TEXT"),
            }
        )

    checks = {
        "100_rows_plus_header": len(xlsx_rows) == 101,
        "exact_header": header == list(HEADER),
        "ranks_1_to_100_once": loaded_ranks == list(range(1, 101)),
        "scores_non_increasing": all(
            loaded_scores[index - 1] >= loaded_scores[index]
            for index in range(1, len(loaded_scores))
        ),
        "row_by_row_value_equality": not mismatches,
        "cell_types_preserved": not type_failures,
    }
    workbook.close()
    return {
        "checks": checks,
        "mismatch_count": len(mismatches),
        "mismatch_rows": mismatches,
        "type_failures": type_failures,
        "preview": preview,
    }
